import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Functions
def workbookName():
    path = './excels/'
    for file in os.listdir(path):
        if (os.path.isfile(os.path.join(path, file)) and ".xlsx" in file):
            xlsx_file = path + file
            return xlsx_file

def getLengthOfArray(a, b):
    for col in range(a, b):
        char = get_column_letter(col)
        i = 1
        while True:
            cell = activeWorksheet[char + str(i)].value
            if cell == None:
                return i
            else:
                i += 1
    return i

def getArrayItems(a, b, end_of_row):
    array = []
    for col in range(a, b):
        char = get_column_letter(col)
        for row in range(2, end_of_row):
            cell = activeWorksheet[char + str(row)].value
            array.append(cell)
    return array

# Loading xlsx
workbookName = workbookName()
loadingWorkbook = load_workbook(workbookName)
activeWorksheet = loadingWorkbook.active

# Getting array of items
appname = workbookName.split("/")[-1]
appname = appname.split(".")[0]
length = getLengthOfArray(3, 4)
entryArray = getArrayItems(3, 4, length)

# MDS query template YAML file
yamlFile = f'''# {datetime.now()}
---
'''
i = 0
while i < len(entryArray):
    yamlEntry = f"""# {i + 1}
-   pkstatus: true
    dataSource: hasura
    querytemplate: '{entryArray[i]}'
    pk:
    appid: {appname}
    paramValuesSchema:
    required:

"""
    yamlFile += yamlEntry
    i += 1

# Creating query template YAML file
with open('./artifacts/mds_querytemplate.txt', 'w') as file:
    file.write(yamlFile)

txt_file = os.path.join('./artifacts/', 'mds_querytemplate.txt')
yaml_file = txt_file.replace('.txt', '.yaml')
os.rename(txt_file, yaml_file)

print('MDS Query Template: File Generation Completed')
