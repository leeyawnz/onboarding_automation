import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Functions
def workbookName():
    path = './excels/'
    for file in os.listdir(path):
        if (os.path.isfile(os.path.join(path, file)) and ".xlsx" in file):
            xlsx_file = path + file
            return xlsx_file

def getLengthOfArray(a, b):
    for col in range(a, b):
        char = get_column_letter(col)
        i = 1
        while True:
            cell = activeWorksheet[char + str(i)].value
            if cell == None:
                return i
            else:
                i += 1
    return i

def getArrayItems(a, b, end_of_row):
    array = []
    for col in range(a, b):
        char = get_column_letter(col)
        for row in range(2, end_of_row):
            cell = activeWorksheet[char + str(row)].value
            array.append(cell)
    return array

# Loading xlsx
workbookName = workbookName()
loadingWorkbook = load_workbook(workbookName)
activeWorksheet = loadingWorkbook.active

# Getting array of items
appname = workbookName.split("/")[-1]
appname = appname.split(".")[0]
length = getLengthOfArray(2, 3)
pureTableArray = getArrayItems(1, 2, length)
editedTableArray = list(filter(lambda item: item is not None, pureTableArray))
editedTableArray = list(map(lambda item: item.lower(), editedTableArray))
fieldsArray = getArrayItems(2, 3, length)

fieldsArrayOfArrays = []
splitFieldsArray = []
i = 0
while i < len(fieldsArray):
    if i == 0:
        splitFieldsArray.append(fieldsArray[i].lower())
        i += 1
    elif pureTableArray[i] == None and pureTableArray[i - 1] == None:
        splitFieldsArray.append(fieldsArray[i].lower())
        i += 1
    elif pureTableArray[i] == None and pureTableArray[i - 1] != None:
        splitFieldsArray.append(fieldsArray[i].lower())
        i += 1
    else:
        fieldsArrayOfArrays.append(splitFieldsArray)
        splitFieldsArray = []
        splitFieldsArray.append(fieldsArray[i].lower())
        i += 1
fieldsArrayOfArrays.append(splitFieldsArray)

# MDS data contract YAML file
yamlFile = f'''# {datetime.now()}
---
role: [{appname}]
filter: {{
    it0001_persg: [],
    it0001_werks: []
}}

table:'''

j = 0
while j < len(editedTableArray):
    fieldsArrayOfArrays[j] = str(fieldsArrayOfArrays[j]).replace("'","")
    yamlEntry = f'''
-   tablename: {editedTableArray[j]}_main
    columns: {fieldsArrayOfArrays[j]}
    limit: null
    allow_aggregations: false
'''
    yamlFile += yamlEntry
    j += 1

# Creating data contract YAML file
with open('./artifacts/mds_datacontract.txt', 'w') as file:
    file.write(yamlFile)

txt_file = os.path.join('./artifacts/', 'mds_datacontract.txt')
yaml_file = txt_file.replace('.txt', '.yaml')
os.rename(txt_file, yaml_file)

print('MDS Data Contract: File Generation Completed')
