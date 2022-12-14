import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# input statements
wb_name = input('Enter Excelsheet Name (.xlsx): ')
data_tb_len = input('Enter length of data (Last cell + 1): ')
data_tb_len = int(data_tb_len)
appname = input('Enter Application Team Name: ')

# loading workbook
wb = load_workbook(wb_name)
ws = wb.active

# full table list including None
full_tb_list = []
for col in range(1, 2):
    char = get_column_letter(col)
    for row in range(2, data_tb_len):
        cell = ws[char + str(row)].value
        full_tb_list.append(cell)

# table list without None
# Need to check which tables have _main and which does not
tb_list = []
for col in range(1, 2):
    char = get_column_letter(col)
    for row in range(2, data_tb_len):
        cell = ws[char + str(row)].value
        if cell != None:
            tb_list.append(cell.lower())

# full field list
full_f_list = []
for col in range(2, 3):
    char = get_column_letter(col)
    for row in range(2, data_tb_len):
        cell = ws[char + str(row)].value
        full_f_list.append(cell.lower())

# splitting full_tb_list to individual arrays
ind_full_tb_list = []
ind_tb_list = []
i = 0
while i < len(full_tb_list):
    if i == 0:
        ind_tb_list.append(full_f_list[i])
        i += 1
    elif full_tb_list[i] == None and full_tb_list[i - 1] != None:
        ind_tb_list.append(full_f_list[i])
        i += 1
    elif full_tb_list[i] == None and full_tb_list[i - 1] == None:
        ind_tb_list.append(full_f_list[i])
        i += 1
    else:
        ind_full_tb_list.append(ind_tb_list)
        ind_tb_list = []
        ind_tb_list.append(full_f_list[i])
        i += 1
ind_full_tb_list.append(ind_tb_list)

# yaml file
with open(f'{appname}.txt', 'a') as file:
    yaml_header = '''
---
role: []
filter: {
    it0001_persg: [],
    it0001_werks: []
}

table:
'''
    file.write(yaml_header)

j = 0
while j < len(tb_list):
    ind_full_tb_list[j] = str(ind_full_tb_list[j]).replace("'", "")
    with open(f'{appname}.txt', 'a') as file:
        yaml_template = f'''
- tablename: {tb_list[j]}
  columns: {ind_full_tb_list[j]}
  limit: null
  allow_aggregations: false
'''
        file.write(yaml_template)
        j += 1

txt_file = os.path.join('.', f'{appname}.txt')
yaml_file = txt_file.replace('.txt', '.yaml')
os.rename(txt_file, yaml_file)
