import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Functions
def workbookName():
    path = './excels/'
    for file in os.listdir(path):
        if (os.path.isfile(os.path.join(path, file)) and ".xlsx" in file):
            xlsx_file = path + file
            print("Detected: " + xlsx_file)
            return xlsx_file

def getLengthOfArray(a, b):
    for col in range(a, b):
        char = get_column_letter(col)
        i = 1
        while True:
            cell = activeWorksheet[char + str(i)].value
            if cell == None:
                return i
            else:
                i += 1
    return i

def getArrayItems(a, b, end_of_row):
    array = []
    for col in range(a, b):
        char = get_column_letter(col)
        for row in range(2, end_of_row):
            cell = activeWorksheet[char + str(row)].value
            array.append(cell)
    return array

def editSFTArray(array):
    cleanArray = []
    for item in array:
        entry = item.split(" ")[0]
        cleanArray.append(entry)
    return cleanArray

def sftEntry(appname, array, httpCallbackArray, snsCallbackArray, i):
    sftEntry = f'''
        {{
            "PutRequest": {{
                "Item": {{
                    "pk": {{
                        "S": "{appname}#{array[i]}"
                    }},
                    "pkstatus": {{
                        "S": "true"
                    }},
                    "httpCallback": {{
                        "S": "{httpCallbackArray[0]}"
                    }},
                    "callback": {{
                        "S": "{snsCallbackArray[0]}"
                    }}
                }}
            }}
        }}'''
    return sftEntry

# Loading xlsx
workbookName = workbookName()
loadingWorkbook = load_workbook(workbookName)
activeWorksheet = loadingWorkbook.active

# Getting array of items
appname = workbookName.split("/")[-1]
appname = appname.split(".")[0]
httpCallback = ""
snsCallback = ""
length = getLengthOfArray(5, 6)
pureEntryArray = getArrayItems(5, 6, length)
editedEntryArray = editSFTArray(pureEntryArray)
httpCallbackEntry = getArrayItems(6, 7, 3)
snsCallbackEntry = getArrayItems(7, 8, 3)

# SFT JSON file
sftFile = '''{
    "sft-authorizer-sit": ['''
sftFooter = '''
    ]
}'''

i = 0
while i < len(editedEntryArray):
    if i != len(editedEntryArray) - 1:
        sftFile += sftEntry(appname, editedEntryArray, httpCallbackEntry, snsCallbackEntry, i) + ','
        i += 1
    else:
        sftFile += sftEntry(appname, editedEntryArray, httpCallbackEntry, snsCallbackEntry, i)
        i += 1
sftFile += sftFooter

# Creating JSON file
with open('./artifacts/sft.txt', 'w') as file:
    file.write(sftFile)

txt_file = os.path.join('./artifacts/', 'sft.txt')
json_file = txt_file.replace('.txt', '.json')
os.rename(txt_file, json_file)
